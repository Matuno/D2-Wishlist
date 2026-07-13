from __future__ import annotations

import argparse
import csv
import html
import itertools
import json
import re
import urllib.error
import sys
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - reported by CLI when needed
    load_workbook = None


SHEET_ID = "1JM-0SlxVDAi-C6rGVlLxa-J1WGewEeL8Qvq4htWZHhY"
SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
BUNGIE_MANIFEST_URL = "https://www.bungie.net/Platform/Destiny2/Manifest/"
TIER_ORDER = ("S", "A", "B", "C", "D", "E", "F")
TIER_INDEX = {tier: index for index, tier in enumerate(TIER_ORDER)}
PPC_VALUES = (0, 1, 2, 3)
BMC_VALUES = (0, 1, 2)
OUTPUT_PREFIX = "AegisWishlist"


@dataclass(frozen=True)
class Plug:
    name: str
    hash: int


@dataclass(frozen=True)
class WishlistRow:
    sheet: str
    name: str
    item_hash: int
    tier: str
    rank: float | None
    notes: str
    barrels: list[Plug] = field(default_factory=list)
    mags: list[Plug] = field(default_factory=list)
    perk1: list[Plug] = field(default_factory=list)
    perk2: list[Plug] = field(default_factory=list)
    origins: list[Plug] = field(default_factory=list)


@dataclass(frozen=True)
class SheetRow:
    sheet: str
    row_number: int
    name: str
    season: str
    tier: str
    rank: float | None
    notes: str
    barrels: list[str]
    mags: list[str]
    perk1: list[str]
    perk2: list[str]
    origins: list[str]


@dataclass(frozen=True)
class AuditIssue:
    severity: str
    sheet: str
    row_number: int
    weapon: str
    field: str
    value: str
    message: str


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("’", "'").replace("‐", "-").replace("‑", "-").replace("–", "-")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"^enhanced\s+", "", text)
    return text


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_plug_cell(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"\r?\n", text)
    cleaned: list[str] = []
    ignored = {"", "none", "n/a", "na", "0"}
    for part in parts:
        item = part.strip()
        if normalize_name(item) in ignored:
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", item):
            continue
        cleaned.append(item)
    return dedupe_preserve_order(cleaned)


def dedupe_preserve_order(values: Iterable[Any]) -> list[Any]:
    seen: set[Any] = set()
    result: list[Any] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def base_weapon_name(name: str) -> str:
    return clean_text(name).splitlines()[0].strip()


def should_include_tier(row_tier: str, min_tier: str) -> bool:
    row_key = clean_text(row_tier).upper()
    min_key = clean_text(min_tier).upper()
    if row_key not in TIER_INDEX or min_key not in TIER_INDEX:
        return False
    return TIER_INDEX[row_key] <= TIER_INDEX[min_key]


def selected_count(selected: set[int], plugs: list[Plug]) -> int:
    return sum(1 for plug in plugs if plug.hash in selected)


def generate_rolls_for_row(row: WishlistRow, ppc: int, bmc: int) -> list[tuple[int, ...]]:
    if ppc not in PPC_VALUES:
        raise ValueError(f"Unsupported PPC value: {ppc}")
    if bmc not in BMC_VALUES:
        raise ValueError(f"Unsupported BMC value: {bmc}")

    ordered_plugs = row.barrels + row.mags + row.perk1 + row.perk2 + row.origins
    ordered_hashes = [plug.hash for plug in ordered_plugs]
    rolls: list[tuple[int, ...]] = []
    seen: set[tuple[int, ...]] = set()

    for size in range(len(ordered_hashes), 0, -1):
        for indexes in itertools.combinations(range(len(ordered_hashes)), size):
            roll = tuple(ordered_hashes[index] for index in indexes)
            if roll in seen:
                continue
            selected = set(roll)
            if not roll_matches_strictness(row, selected, ppc, bmc):
                continue
            seen.add(roll)
            rolls.append(roll)
    return rolls


def roll_matches_strictness(row: WishlistRow, selected: set[int], ppc: int, bmc: int) -> bool:
    barrel_count = selected_count(selected, row.barrels)
    mag_count = selected_count(selected, row.mags)
    perk1_count = selected_count(selected, row.perk1)
    perk2_count = selected_count(selected, row.perk2)
    origin_count = selected_count(selected, row.origins)

    if origin_count == len(selected):
        return False

    if ppc > 0:
        required_p1 = min(ppc, len(row.perk1))
        required_p2 = min(ppc, len(row.perk2))
        if perk1_count < required_p1 or perk2_count < required_p2:
            return False

    if bmc == 1 and barrel_count + mag_count < 1:
        return False
    if bmc == 2 and (barrel_count < 1 or mag_count < 1):
        return False
    return True


def generate_wishlist_text(rows: list[WishlistRow], min_tier: str, ppc: int, bmc: int) -> str:
    lines = [
        f"title:Aegis Wishlist MR{min_tier}_PPC{ppc}_BMC{bmc}",
        (
            "description:Generated from Aegis spreadsheet. "
            f"Filters: Grade>={min_tier}, PPC={ppc}, BMC={bmc}"
        ),
        "",
    ]

    for row in rows:
        if not should_include_tier(row.tier, min_tier):
            continue
        rolls = generate_rolls_for_row(row, ppc=ppc, bmc=bmc)
        if not rolls:
            continue
        lines.append(f"// {one_line(row.name)}")
        note = build_note(row)
        if note:
            lines.append(f"//notes:{note}")
        for roll in rolls:
            perks = ",".join(str(hash_value) for hash_value in roll)
            lines.append(f"dimwishlist:item={row.item_hash}&perks={perks}")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def one_line(value: str) -> str:
    return re.sub(r"\s+", " ", clean_text(value))


def build_note(row: WishlistRow) -> str:
    parts = [f"{row.tier}-Tier"]
    if row.rank is not None:
        parts.append(f"Rank {format_rank(row.rank)}")
    if row.notes:
        parts.append(one_line(row.notes))
    if row.origins:
        parts.append("Origin: " + ", ".join(plug.name for plug in row.origins))
    return ". ".join(part for part in parts if part).replace("|", "/")


def format_rank(rank: float) -> str:
    return str(int(rank)) if float(rank).is_integer() else str(rank)


def output_file_name(min_tier: str, ppc: int, bmc: int) -> str:
    return f"{OUTPUT_PREFIX}_MR{min_tier}_PPC{ppc}_BMC{bmc}.txt"


def build_index(out_dir: Path, raw_base_url: str = "") -> dict[str, Any]:
    files: list[dict[str, Any]] = []
    base = raw_base_url.rstrip("/") or f"https://raw.githubusercontent.com/<user>/<repo>/main/{out_dir.name}"
    for min_tier in TIER_ORDER:
        for ppc in PPC_VALUES:
            for bmc in BMC_VALUES:
                file_name = output_file_name(min_tier, ppc, bmc)
                files.append(
                    {
                        "minimumRank": min_tier,
                        "ppc": ppc,
                        "bmc": bmc,
                        "fileName": file_name,
                        "localPath": f"{out_dir.name}/{file_name}",
                        "rawUrl": f"{base}/{file_name}",
                    }
                )
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "outputDirectory": out_dir.name,
        "files": files,
    }


def write_outputs(
    rows: list[WishlistRow],
    out_dir: Path,
    raw_base_url: str = "",
    audit_issues: list[AuditIssue] | None = None,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    for min_tier in TIER_ORDER:
        for ppc in PPC_VALUES:
            for bmc in BMC_VALUES:
                file_name = output_file_name(min_tier, ppc, bmc)
                (out_dir / file_name).write_text(
                    generate_wishlist_text(rows, min_tier=min_tier, ppc=ppc, bmc=bmc),
                    encoding="utf-8",
                )

    index = build_index(out_dir, raw_base_url=raw_base_url)
    if metadata:
        index["metadata"] = metadata
    (out_dir / "index.json").write_text(json.dumps(index, indent=2), encoding="utf-8")
    (out_dir / "index.html").write_text(render_configurator_html(index), encoding="utf-8")
    write_audit(out_dir / "audit.csv", audit_issues or [])
    return index


def write_audit(path: Path, audit_issues: list[AuditIssue]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("severity", "sheet", "row_number", "weapon", "field", "value", "message"),
        )
        writer.writeheader()
        for issue in audit_issues:
            writer.writerow(
                {
                    "severity": issue.severity,
                    "sheet": issue.sheet,
                    "row_number": issue.row_number,
                    "weapon": issue.weapon,
                    "field": issue.field,
                    "value": issue.value,
                    "message": issue.message,
                }
            )


def render_configurator_html(index: dict[str, Any]) -> str:
    index_json = json.dumps(index, separators=(",", ":"))
    css = """
*{box-sizing:border-box}body{margin:0;font-family:Segoe UI,Arial,sans-serif;background:#f6f7f9;color:#20242b}main{max-width:920px;margin:0 auto;padding:32px 20px 48px}h1{font-size:30px;margin:0 0 8px}p{color:#53606f;line-height:1.5;margin:0 0 24px}.toolbar{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:16px;margin:24px 0}.field{display:flex;flex-direction:column;gap:8px}label{font-weight:600;font-size:14px}select,input{height:42px;border:1px solid #c7ced8;border-radius:6px;background:#fff;padding:0 12px;font-size:15px}.result{border:1px solid #c7ced8;background:#fff;border-radius:8px;padding:18px;display:grid;gap:12px}.row{display:grid;grid-template-columns:140px 1fr;gap:16px;align-items:center}.key{font-weight:600;color:#53606f}.value{font-family:Consolas,Menlo,monospace;overflow-wrap:anywhere}.copyline{display:flex;gap:8px}.copyline input{flex:1}.copyline button{height:42px;border:1px solid #20242b;background:#20242b;color:#fff;border-radius:6px;padding:0 14px;font-weight:600;cursor:pointer}.hint{font-size:13px;color:#6c7684;margin-top:18px}@media (max-width:720px){.toolbar{grid-template-columns:1fr}.row{grid-template-columns:1fr;gap:4px}.copyline{flex-direction:column}.copyline button{width:100%}}
"""
    script = """
const DATA = __INDEX__;
const rank = document.querySelector("#rank");
const ppc = document.querySelector("#ppc");
const bmc = document.querySelector("#bmc");
const fileName = document.querySelector("#fileName");
const localPath = document.querySelector("#localPath");
const rawUrl = document.querySelector("#rawUrl");
const copy = document.querySelector("#copy");
function update(){
  const entry = DATA.files.find(item => item.minimumRank === rank.value && item.ppc === Number(ppc.value) && item.bmc === Number(bmc.value));
  fileName.textContent = entry.fileName;
  localPath.textContent = entry.localPath;
  rawUrl.value = entry.rawUrl;
}
for (const node of [rank, ppc, bmc]) node.addEventListener("change", update);
copy.addEventListener("click", async () => {
  await navigator.clipboard.writeText(rawUrl.value);
  copy.textContent = "Copied";
  setTimeout(() => copy.textContent = "Copy", 1200);
});
update();
""".replace("__INDEX__", index_json)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DIM Aegis Wishlist Configurator</title>
  <style>{css}</style>
</head>
<body>
  <main>
    <h1>DIM Aegis Wishlist Configurator</h1>
    <p>Select minimum rank, main perk strictness, and barrel/mag strictness. Use the raw URL after hosting these files on GitHub.</p>
    <section class="toolbar" aria-label="Wishlist filters">
      <div class="field">
        <label for="rank">Minimum rank</label>
        <select id="rank">{options(TIER_ORDER)}</select>
      </div>
      <div class="field">
        <label for="ppc">Main perks</label>
        <select id="ppc">{options(PPC_VALUES, "PPC")}</select>
      </div>
      <div class="field">
        <label for="bmc">Barrel / mag</label>
        <select id="bmc">{options(BMC_VALUES, "BMC")}</select>
      </div>
    </section>
    <section class="result" aria-label="Selected wishlist">
      <div class="row"><div class="key">File</div><div class="value" id="fileName"></div></div>
      <div class="row"><div class="key">Local path</div><div class="value" id="localPath"></div></div>
      <div class="row"><div class="key">Raw URL</div><div class="copyline"><input id="rawUrl" readonly><button id="copy" type="button">Copy</button></div></div>
    </section>
    <p class="hint">BMC0 keeps barrel and magazine optional. BMC1 requires either one. BMC2 requires both.</p>
  </main>
  <script>{script}</script>
</body>
</html>
"""


def options(values: Iterable[Any], prefix: str = "") -> str:
    return "".join(
        f'<option value="{html.escape(str(value))}">{html.escape(prefix + str(value))}</option>'
        for value in values
    )


class ManifestResolver:
    def __init__(
        self,
        item_defs: dict[str, dict[str, Any]],
        plug_set_defs: dict[str, dict[str, Any]],
        overrides: dict[str, Any] | None = None,
    ) -> None:
        self.item_defs = item_defs
        self.plug_set_defs = plug_set_defs
        self.overrides = overrides or {}
        self.name_to_weapon_hashes: dict[str, list[int]] = {}
        self.plug_name_by_hash: dict[int, str] = {}
        for hash_key, definition in item_defs.items():
            display = definition.get("displayProperties") or {}
            name = display.get("name") or ""
            try:
                hash_value = int(hash_key)
            except ValueError:
                hash_value = int(definition.get("hash", 0))
            if name:
                self.plug_name_by_hash[hash_value] = name
            if definition.get("itemType") == 3 and name:
                self.name_to_weapon_hashes.setdefault(normalize_name(name), []).append(hash_value)

    def resolve(self, sheet_row: SheetRow) -> tuple[WishlistRow | None, list[AuditIssue]]:
        issues: list[AuditIssue] = []
        item_hash = self.resolve_item_hash(sheet_row, issues)
        if item_hash is None:
            return None, issues

        option_hashes = self.item_socket_hashes(item_hash)
        resolved_groups: dict[str, list[Plug]] = {}
        for field, names in {
            "Barrel": sheet_row.barrels,
            "Mag": sheet_row.mags,
            "Perk 1": sheet_row.perk1,
            "Perk 2": sheet_row.perk2,
            "Origin Trait": sheet_row.origins,
        }.items():
            resolved_groups[field] = self.resolve_plug_group(sheet_row, field, names, option_hashes, issues)

        return (
            WishlistRow(
                sheet=sheet_row.sheet,
                name=sheet_row.name,
                item_hash=item_hash,
                tier=sheet_row.tier,
                rank=sheet_row.rank,
                notes=sheet_row.notes,
                barrels=resolved_groups["Barrel"],
                mags=resolved_groups["Mag"],
                perk1=resolved_groups["Perk 1"],
                perk2=resolved_groups["Perk 2"],
                origins=resolved_groups["Origin Trait"],
            ),
            issues,
        )

    def resolve_item_hash(self, sheet_row: SheetRow, issues: list[AuditIssue]) -> int | None:
        item_overrides = self.overrides.get("items", {})
        for key in (sheet_row.name, base_weapon_name(sheet_row.name)):
            if key in item_overrides:
                return int(item_overrides[key])

        candidates = self.name_to_weapon_hashes.get(normalize_name(base_weapon_name(sheet_row.name)), [])
        if not candidates:
            issues.append(issue(sheet_row, "error", "Name", sheet_row.name, "No manifest weapon matched this name"))
            return None

        desired_names = (
            sheet_row.barrels + sheet_row.mags + sheet_row.perk1 + sheet_row.perk2 + sheet_row.origins
        )
        scored: list[tuple[int, int]] = []
        for candidate in candidates:
            option_hashes = self.item_socket_hashes(candidate)
            option_names = {normalize_name(self.plug_name_by_hash.get(hash_value, "")) for hash_value in option_hashes}
            score = sum(1 for name in desired_names if normalize_name(name) in option_names)
            scored.append((score, candidate))
        scored.sort(key=lambda item: (-item[0], item[1]))
        best_score = scored[0][0]
        best_candidates = [candidate for score, candidate in scored if score == best_score]
        if best_score == 0:
            issues.append(
                issue(
                    sheet_row,
                    "error",
                    "Name",
                    sheet_row.name,
                    "No matching candidate exposed the desired socket plugs; add config/overrides.yml",
                )
            )
            return None
        if len(best_candidates) > 1:
            selected = sorted(best_candidates, key=self.candidate_sort_key, reverse=True)[0]
            issues.append(
                issue(
                    sheet_row,
                    "warning",
                    "Name",
                    sheet_row.name,
                    f"Multiple manifest weapons matched; selected {selected} by release and collectible tie-breakers",
                )
            )
            return selected
        return best_candidates[0]

    def candidate_sort_key(self, item_hash: int) -> tuple[int, int, int]:
        definition = self.item_defs.get(str(item_hash), {})
        release = 0
        for trait_id in definition.get("traitIds") or []:
            match = re.search(r"releases\.v(\d+)", trait_id)
            if match:
                release = max(release, int(match.group(1)))
        has_collectible = 1 if definition.get("collectibleHash") else 0
        return (release, has_collectible, item_hash)

    def item_socket_hashes(self, item_hash: int) -> set[int]:
        definition = self.item_defs.get(str(item_hash), {})
        hashes: set[int] = set()
        for socket in (definition.get("sockets") or {}).get("socketEntries", []):
            single = socket.get("singleInitialItemHash")
            if single:
                hashes.add(int(single))
            for entry in socket.get("reusablePlugItems", []):
                plug_hash = entry.get("plugItemHash")
                if plug_hash:
                    hashes.add(int(plug_hash))
            for key in ("reusablePlugSetHash", "randomizedPlugSetHash"):
                plug_set_hash = socket.get(key)
                if plug_set_hash:
                    plug_set = self.plug_set_defs.get(str(plug_set_hash), {})
                    for entry in plug_set.get("reusablePlugItems", []):
                        plug_hash = entry.get("plugItemHash")
                        if plug_hash:
                            hashes.add(int(plug_hash))
        return hashes

    def resolve_plug_group(
        self,
        sheet_row: SheetRow,
        field_name: str,
        names: list[str],
        option_hashes: set[int],
        issues: list[AuditIssue],
    ) -> list[Plug]:
        plug_overrides = (self.overrides.get("plugs", {}).get(sheet_row.name, {}) or {}) | (
            self.overrides.get("plugs", {}).get(base_weapon_name(sheet_row.name), {}) or {}
        )
        ignored_plugs = self.ignored_plugs(sheet_row, field_name)
        result: list[Plug] = []
        for name in names:
            if normalize_name(name) in ignored_plugs:
                issues.append(
                    issue(
                        sheet_row,
                        "warning",
                        field_name,
                        name,
                        "Plug ignored by config/overrides.yml because it is not available on the matched weapon",
                    )
                )
                continue
            if name in plug_overrides:
                result.append(Plug(name, int(plug_overrides[name])))
                continue
            matches = [
                hash_value
                for hash_value in option_hashes
                if normalize_name(self.plug_name_by_hash.get(hash_value, "")) == normalize_name(name)
            ]
            matches = sorted(set(matches))
            if not matches:
                severity = "error" if field_name in {"Perk 1", "Perk 2"} else "warning"
                issues.append(issue(sheet_row, severity, field_name, name, "Plug was not available on matched weapon"))
                continue
            exact = [
                hash_value
                for hash_value in matches
                if clean_text(self.plug_name_by_hash.get(hash_value, "")) == clean_text(name)
            ]
            result.append(Plug(name, exact[0] if exact else matches[0]))
        return dedupe_plugs(result)

    def ignored_plugs(self, sheet_row: SheetRow, field_name: str) -> set[str]:
        ignored_config = self.overrides.get("ignored_plugs", {})
        values: list[str] = []
        for key in (sheet_row.name, base_weapon_name(sheet_row.name)):
            entry = ignored_config.get(key)
            if not entry:
                continue
            if isinstance(entry, dict):
                for nested_key in (field_name, "*"):
                    nested = entry.get(nested_key)
                    if isinstance(nested, list):
                        values.extend(clean_text(value) for value in nested)
                    elif nested:
                        values.append(clean_text(nested))
            elif isinstance(entry, list):
                values.extend(clean_text(value) for value in entry)
            else:
                values.append(clean_text(entry))
        return {normalize_name(value) for value in values if value}


def dedupe_plugs(plugs: list[Plug]) -> list[Plug]:
    seen: set[int] = set()
    result: list[Plug] = []
    for plug in plugs:
        if plug.hash not in seen:
            seen.add(plug.hash)
            result.append(plug)
    return result


def issue(sheet_row: SheetRow, severity: str, field: str, value: str, message: str) -> AuditIssue:
    return AuditIssue(
        severity=severity,
        sheet=sheet_row.sheet,
        row_number=sheet_row.row_number,
        weapon=sheet_row.name,
        field=field,
        value=value,
        message=message,
    )


def parse_workbook_bytes(workbook_bytes: bytes) -> list[SheetRow]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse the Aegis spreadsheet")
    from io import BytesIO

    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    rows: list[SheetRow] = []
    for sheet in workbook.worksheets:
        header_values = next(
            sheet.iter_rows(min_row=2, max_row=2, max_col=min(sheet.max_column or 30, 30), values_only=True),
            None,
        )
        if not header_values:
            continue
        headers = [clean_text(value) for value in header_values]
        required = ["Name", "Barrel", "Mag", "Perk 1", "Perk 2", "Origin Trait", "Notes", "Rank", "Tier"]
        if not all(header in headers for header in required):
            continue
        indexes = {header: headers.index(header) for header in required if header in headers}
        indexes["Season"] = headers.index("Season") if "Season" in headers else -1
        max_col = max(indexes.values()) + 1
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=3, max_col=max_col, values_only=True),
            start=3,
        ):
            name = clean_text(values[indexes["Name"]])
            tier = clean_text(values[indexes["Tier"]]).upper()
            if not name or tier not in TIER_INDEX:
                continue
            perk1 = split_plug_cell(values[indexes["Perk 1"]])
            perk2 = split_plug_cell(values[indexes["Perk 2"]])
            if not perk1 and not perk2:
                continue
            rows.append(
                SheetRow(
                    sheet=sheet.title,
                    row_number=row_number,
                    name=name,
                    season=clean_text(values[indexes["Season"]]) if indexes["Season"] >= 0 else "",
                    tier=tier,
                    rank=parse_rank(values[indexes["Rank"]]),
                    notes=clean_text(values[indexes["Notes"]]),
                    barrels=split_plug_cell(values[indexes["Barrel"]]),
                    mags=split_plug_cell(values[indexes["Mag"]]),
                    perk1=perk1,
                    perk2=perk2,
                    origins=split_plug_cell(values[indexes["Origin Trait"]]),
                )
            )
    return rows


def parse_rank(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


TRANSIENT_HTTP_CODES = {429, 500, 502, 503, 504}


def read_url(url: str, retries: int = 3) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "D2-Wishlist-Generator/1.0"})
    for attempt in range(1, retries + 1):
        try:
            with urllib.request.urlopen(request, timeout=120) as response:
                return response.read()
        except urllib.error.HTTPError as exc:
            if exc.code not in TRANSIENT_HTTP_CODES or attempt == retries:
                raise
        except urllib.error.URLError:
            if attempt == retries:
                raise
    raise RuntimeError(f"Failed to fetch {url}")


def fetch_json(url: str) -> dict[str, Any]:
    return json.loads(read_url(url).decode("utf-8"))


def fetch_bytes(url: str) -> bytes:
    return read_url(url)


def load_bungie_manifest() -> tuple[str, dict[str, Any], dict[str, Any]]:
    manifest = fetch_json(BUNGIE_MANIFEST_URL)["Response"]
    paths = manifest["jsonWorldComponentContentPaths"]["en"]
    item_defs = fetch_json("https://www.bungie.net" + paths["DestinyInventoryItemDefinition"])
    plug_set_defs = fetch_json("https://www.bungie.net" + paths["DestinyPlugSetDefinition"])
    return manifest["version"], item_defs, plug_set_defs


def load_overrides(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover - only hit without dependency
        raise RuntimeError("PyYAML is required when config/overrides.yml exists") from exc
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def resolve_rows(sheet_rows: list[SheetRow], resolver: ManifestResolver) -> tuple[list[WishlistRow], list[AuditIssue]]:
    resolved: list[WishlistRow] = []
    issues: list[AuditIssue] = []
    for sheet_row in sheet_rows:
        row, row_issues = resolver.resolve(sheet_row)
        issues.extend(row_issues)
        if row and not any(problem.severity == "error" for problem in row_issues):
            resolved.append(row)
    return resolved, issues


def run_generation(
    out_dir: Path,
    raw_base_url: str = "",
    sheet_url: str = SHEET_EXPORT_URL,
    overrides_path: Path = Path("config/overrides.yml"),
    allow_unresolved: bool = False,
) -> int:
    sheet_bytes = fetch_bytes(sheet_url)
    manifest_version, item_defs, plug_set_defs = load_bungie_manifest()
    sheet_rows = parse_workbook_bytes(sheet_bytes)
    resolver = ManifestResolver(item_defs, plug_set_defs, overrides=load_overrides(overrides_path))
    rows, issues = resolve_rows(sheet_rows, resolver)
    metadata = {
        "sheetUrl": sheet_url,
        "manifestVersion": manifest_version,
        "sheetRows": len(sheet_rows),
        "resolvedRows": len(rows),
        "errors": sum(1 for issue_item in issues if issue_item.severity == "error"),
    }
    errors = [item for item in issues if item.severity == "error"]
    if errors and not allow_unresolved:
        out_dir.mkdir(parents=True, exist_ok=True)
        write_audit(out_dir / "audit.csv", issues)
        print(
            f"Found {len(errors)} unresolved mappings. See {out_dir / 'audit.csv'}; "
            "add config/overrides.yml or rerun with --allow-unresolved.",
            file=sys.stderr,
        )
        return 2
    write_outputs(rows, out_dir=out_dir, raw_base_url=raw_base_url, audit_issues=issues, metadata=metadata)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate DIM Aegis wishlist files with BMC strictness.")
    subparsers = parser.add_subparsers(dest="command")
    generate = subparsers.add_parser("generate", help="Generate dist wishlist files and configurator.")
    generate.add_argument("--out", default="dist", type=Path, help="Output directory.")
    generate.add_argument("--raw-base-url", default="", help="Base GitHub raw URL for generated files.")
    generate.add_argument("--sheet-url", default=SHEET_EXPORT_URL, help="Aegis spreadsheet .xlsx export URL.")
    generate.add_argument("--overrides", default=Path("config/overrides.yml"), type=Path, help="Override YAML path.")
    generate.add_argument("--allow-unresolved", action="store_true", help="Write partial outputs despite audit errors.")
    dim_links = subparsers.add_parser(
        "monument-dim-links",
        help="Generate long-form DIM links for Monument of Triumph build docs.",
    )
    dim_links.add_argument(
        "--build-root",
        default=Path("builds/monument-of-triumph"),
        type=Path,
        help="Build docs root.",
    )
    dim_links.add_argument("--season-number", default=29, type=int, help="Artifact season number for DIM.")
    dim_links.add_argument(
        "--allow-unresolved",
        action="store_true",
        help="Write links even when a referenced build item or perk is unresolved.",
    )

    args = parser.parse_args(argv)
    if args.command == "generate":
        return run_generation(
            out_dir=args.out,
            raw_base_url=args.raw_base_url,
            sheet_url=args.sheet_url,
            overrides_path=args.overrides,
            allow_unresolved=args.allow_unresolved,
        )
    if args.command == "monument-dim-links":
        from .dim_loadouts import generate_monument_dim_links

        manifest_version, item_defs, plug_set_defs = load_bungie_manifest()
        results = generate_monument_dim_links(
            build_root=args.build_root,
            item_defs=item_defs,
            plug_set_defs=plug_set_defs,
            season_number=args.season_number,
        )
        unresolved = [
            f"{result.doc.path}: {issue}"
            for result in results
            for issue in result.unresolved
        ]
        if unresolved:
            for issue in unresolved:
                print(issue, file=sys.stderr)
            if not args.allow_unresolved:
                print(
                    f"Generated links with manifest {manifest_version}, but found {len(unresolved)} unresolved references.",
                    file=sys.stderr,
                )
                return 2
        print(f"Generated {len(results)} long-form DIM links with manifest {manifest_version}.")
        return 0
    parser.print_help()
    return 1
